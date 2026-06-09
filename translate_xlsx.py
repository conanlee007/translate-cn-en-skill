#!/usr/bin/env python3
"""
XLSX Chinese → English translator with full format preservation.
Covers: all sheets, all cells (skips formulas and pure numbers).
Merged cells: master cell translated, slave cells untouched.

Usage: python translate_xlsx.py <input.xlsx> [output.xlsx]
Requires: ANTHROPIC_API_KEY environment variable
"""
import json
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import anthropic
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

CLAUDE_MODEL = "claude-sonnet-4-6"
BATCH_SIZE = 20
MAX_WORKERS = 5

CJK_RE = re.compile(r'[一-鿿㐀-䶿＀-￯　-〿⺀-⻿]')


def has_cjk(text: str) -> bool:
    return bool(CJK_RE.search(text))


# ── Claude translation (batch) ────────────────────────────────────────────────

def _parse_response(raw: str, expected_len: int, fallback: list[str]) -> list[str]:
    """
    Multi-strategy JSON parser — never raises, always returns a list of expected_len.
    Strategy 1: direct json.loads on extracted [...] block
    Strategy 2: regex extract all quoted strings
    Strategy 3: return fallback (original texts)
    """
    raw = re.sub(r'^```(?:json)?\s*', '', raw.strip())
    raw = re.sub(r'\s*```\s*$', '', raw)

    m = re.search(r'\[.*\]', raw, re.DOTALL)
    if m:
        try:
            parsed = json.loads(m.group())
            if isinstance(parsed, list):
                result = [str(x) for x in parsed]
                if len(result) < expected_len:
                    result += fallback[len(result):]
                return result[:expected_len]
        except Exception:
            pass

    strings = re.findall(r'"((?:[^"\\]|\\.)*)"', raw)
    if strings:
        strings = [s.replace('\\"', '"').replace('\\n', '\n').replace('\\\\', '\\')
                   for s in strings]
        if len(strings) < expected_len:
            strings += fallback[len(strings):]
        return strings[:expected_len]

    return fallback


def _api_call_with_backoff(client, texts: list[str]) -> str:
    """Call Claude API with exponential backoff on rate limit / server errors."""
    prompt = (
        "Translate the following Chinese texts to English. "
        "Return ONLY a JSON array with exactly the same number of elements in the same order.\n"
        "Rules:\n"
        "- Use standard English financial/accounting terminology\n"
        "- Express Chinese yuan as CNY (not RMB)\n"
        "- Keep numbers, percentages, dates, stock codes, and '--' exactly unchanged\n"
        "- Keep empty strings as empty strings\n"
        "- No explanations, no markdown fences — only the JSON array\n\n"
        + json.dumps(texts, ensure_ascii=False)
    )
    for attempt in range(5):
        try:
            msg = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=8096,
                messages=[{"role": "user", "content": prompt}],
            )
            return msg.content[0].text
        except Exception as e:
            err = str(e).lower()
            if 'rate' in err or '429' in err or 'overload' in err:
                wait = 2 ** attempt * 10
                print(f' [rate-limit, wait {wait}s]', end='', flush=True)
                time.sleep(wait)
            elif attempt < 4:
                time.sleep(3)
            else:
                raise
    raise RuntimeError("API call failed after 5 attempts")


def translate_batch(client, texts: list[str]) -> list[str]:
    """Translate a batch with progressive fallback: full → halves → one-by-one."""
    try:
        raw = _api_call_with_backoff(client, texts)
        return _parse_response(raw, len(texts), texts)
    except Exception:
        pass

    print(' [½-batch]', end='', flush=True)
    results = []
    mid = max(len(texts) // 2, 1)
    for chunk in [texts[:mid], texts[mid:]]:
        if not chunk:
            continue
        try:
            raw = _api_call_with_backoff(client, chunk)
            results.extend(_parse_response(raw, len(chunk), chunk))
        except Exception:
            print(' [1-by-1]', end='', flush=True)
            for t in chunk:
                try:
                    raw = _api_call_with_backoff(client, [t])
                    results.extend(_parse_response(raw, 1, [t]))
                except Exception:
                    results.append(t)
    return results


def translate_all(client, texts: list[str], checkpoint_path: str) -> list[str]:
    """Parallel batch translation with per-batch checkpointing."""
    total = len(texts)
    batches = [(idx, texts[i:i + BATCH_SIZE])
               for idx, i in enumerate(range(0, total, BATCH_SIZE))]
    n_batches = len(batches)

    completed: dict[int, list[str]] = {}
    if os.path.exists(checkpoint_path):
        try:
            with open(checkpoint_path, 'r', encoding='utf-8') as f:
                ckpt = json.load(f)
            if 'batches' in ckpt:
                completed = {int(k): v for k, v in ckpt['batches'].items()}
            done_texts = sum(len(v) for v in completed.values())
            print(f"    ↩ Resuming: {len(completed)}/{n_batches} batches done "
                  f"({done_texts}/{total} texts)")
        except Exception:
            completed = {}

    lock = threading.Lock()
    done_count = sum(len(v) for v in completed.values())

    def progress_line(done: int, batches_done: int) -> str:
        pct = done / total * 100
        bar_len = 30
        filled = int(bar_len * done / total)
        bar = '█' * filled + '░' * (bar_len - filled)
        return (f"  [{bar}] {done:,}/{total:,} texts "
                f"({pct:.1f}%) | batch {batches_done}/{n_batches}")

    def process_batch(idx: int, batch: list[str]) -> None:
        nonlocal done_count
        result = translate_batch(client, batch)
        with lock:
            completed[idx] = result
            done_count += len(result)
            batches_done = len(completed)
            print(progress_line(done_count, batches_done), flush=True)
            try:
                with open(checkpoint_path, 'w', encoding='utf-8') as f:
                    json.dump(
                        {'batches': {str(k): v for k, v in completed.items()}},
                        f, ensure_ascii=False
                    )
            except Exception as e:
                print(f'  [ckpt err: {e}]', flush=True)

    pending = [(idx, batch) for idx, batch in batches if idx not in completed]
    print(f"    {len(pending)} batches pending, {MAX_WORKERS} parallel workers")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_batch, idx, batch): idx
                   for idx, batch in pending}
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"\n    [batch {futures[future]} error: {e}]")

    results = []
    for idx, batch in batches:
        results.extend(completed.get(idx, batch))

    try:
        os.remove(checkpoint_path)
    except Exception:
        pass

    return results


# ── XLSX extraction & writing ─────────────────────────────────────────────────

def collect_xlsx_texts(wb):
    """
    Collect all cells with CJK text across all sheets.
    Skips: MergedCell slaves, formulas (=...), non-string values without CJK.
    Returns: (refs: list[Cell], texts: list[str])
    """
    refs, texts = [], []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # Skip merged slave cells — they have no editable value
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                # Skip formulas
                if isinstance(val, str) and val.startswith('='):
                    continue
                text = str(val)
                if has_cjk(text):
                    refs.append(cell)
                    texts.append(text)
    return refs, texts


# ── Main ──────────────────────────────────────────────────────────────────────

def translate_xlsx(input_path: str, output_path: str):
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        sys.exit('Error: ANTHROPIC_API_KEY is not set.')

    client = anthropic.Anthropic(api_key=api_key)

    print(f'Loading workbook: {input_path}')
    # keep_vba=False is fine for .xlsx; use keep_vba=True for .xlsm if needed
    wb = load_workbook(input_path)

    sheet_names = wb.sheetnames
    print(f'  Sheets: {", ".join(sheet_names)}')

    print('Collecting cells with Chinese text...')
    refs, texts = collect_xlsx_texts(wb)
    print(f'  {len(texts)} cells with Chinese text across {len(sheet_names)} sheet(s)')

    if not texts:
        print('Nothing to translate.')
        wb.save(output_path)
        print(f'Saved → {output_path}')
        return

    checkpoint = output_path + '.checkpoint.json'
    print(f'\nTranslating {len(texts)} cells...')
    translated = translate_all(client, texts, checkpoint)

    print('\nWriting translations back...')
    for cell, trans in zip(refs, translated):
        cell.value = trans

    wb.save(output_path)
    print(f'\nSaved → {output_path}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python translate_xlsx.py <input.xlsx> [output.xlsx]')
        sys.exit(1)
    inp = sys.argv[1]
    out = (sys.argv[2] if len(sys.argv) > 2
           else str(Path(inp).parent / (Path(inp).stem + '-EN.xlsx')))
    translate_xlsx(inp, out)
